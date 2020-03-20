# import configuration  # .py file that I store my username/password/token/server domain
import sys
import os
import copy
import datetime
import math

# for Jira control
from jira import JIRA
from jira.exceptions import JIRAError
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing import Drawing


DevTracker = 'http://hlm.lge.com/issue'
QTracker = 'http://hlm.lge.com/qi'
ID = # insert ID
PASSWORD = # insert PW
FILENAME = "D:/test/newGantt.xlsx"

# jira = JIRA(DevTracker, basic_auth=(ID, PASSWORD))

def labelCheck(label) :
    priority = ""
    for l in label :
        if(l == "PMO.P1") : priority = "P1"
        elif(l == "PMO.P2") : priority = "P2"
    return priority

def componCheck(compon) :
    scm = ""
    for l in compon :
        if(l.name == "_SCM") : scm = "O"
    return scm

def nowSprintConvertor():
    start = "2019-09-02"
    startDlist = start.split("-")

    startdate = datetime.date(int(startDlist[0]), int(
        startDlist[1]), int(startDlist[2]))
    milestoneDate = datetime.date.today()
    deltaDate = milestoneDate - startdate
    
    if(deltaDate.days > 0):
        sprint = math.floor(deltaDate.days/14) + 1
        print(sprint)
    else:
        sprint = 0

    return sprint


def sprintConvertor(date):
    start = "2019-09-02"
    startDlist = start.split("-")

    startdate = datetime.date(int(startDlist[0]), int(
        startDlist[1]), int(startDlist[2]))
    dateDlist = date.split("-")
    milestoneDate = datetime.date(
        int(dateDlist[0]), int(dateDlist[1]), int(dateDlist[2]))
    deltaDate = milestoneDate - startdate
    if(deltaDate.days > 0):
        sprint = math.floor(deltaDate.days/14) + 1
    else:
        sprint = 0

    return sprint


def archiListUp(initiative_count, initiative, fixver, row, startSP, sheet1, prior, scm):
    jqlfilter = "issuetype = EPIC and component = _architecture AND issuefunction in linkedIssuesOf(\"key=" + \
        initiative+"\")"
    issue_search_convert = jira.search_issues(jqlfilter, maxResults=100)
    archi_total = issue_search_convert.total
    print("architotal : {}".format(archi_total))
    archi_count = 1
    if(archi_total != 0):
        startpt = startSP + 12
        for key in issue_search_convert:
            archi = jira.issue(key)
            if(archi.fields.issuetype.name == "Epic"):
                archi_type = "Archi"
            else:
                break
            archi_key = archi.key
            archi_summary = "Architectural Review"
            archi_assignee_row = archi.fields.assignee.displayName
            a_assignee_temp_list = archi_assignee_row.split(" ")
            archi_assignee = a_assignee_temp_list[0]
            archi_assignee_id = archi.fields.assignee.name
            archi_status = archi.fields.status.name

            archi_due = archi.fields.duedate
            if(archi_due == None):
                archi_release = 1
                asp = "none"
            else:
                archi_release = sprintConvertor(archi_due)
                asp = archi_release

            aSpColumn = archi_release + 12
            sheet1.cell(row=row, column=1).value = str(
                initiative_count) + "-A" + str(archi_count)
            sheet1.cell(row=row, column=2).value = archi_type
            sheet1.cell(row=row, column=3).value = archi_key
            sheet1.cell(row=row, column=4).value = archi_summary
            sheet1.cell(row=row, column=5).value = archi_assignee
            sheet1.cell(row=row, column=6).value = archi_status
            sheet1.cell(row=row, column=7).value = archi_assignee_id
            sheet1.cell(row=row, column=8).value = "=VLOOKUP(" + \
                sheet1.cell(row=row, column=7).coordinate+",org!$A:$D,4,0)"
            sheet1.cell(
                row=row, column=3).hyperlink = "http://hlm.lge.com/issue/browse/"+archi_key
            sheet1.cell(row=row, column=aSpColumn).value = asp
            sheet1.cell(row=row, column=9).value = fixver
            sheet1.cell(row=row, column=10).value = prior
            sheet1.cell(row=row, column=11).value = scm

            for j in range(startpt, aSpColumn+1):
                sheet1.cell(row=row, column=j).fill = PatternFill(
                    fill_type='solid', fgColor="A9A9A9")

            archi_count = archi_count + 1
            print("row:{}, type:{}, key: {}".format(row, archi_type, archi_key))
            row = row + 1

    return row, archi_total


def milestoneListUp(initiative_count, initiative, fixver, row, startSP, sheet1, archiTotal,prior, scm):
    jqlfilter = "type=milestone and issueFunction in linkedIssuesOfRecursiveLimited(\"key=" + \
        initiative+"\", 2) ORDER BY due ASC"
    issue_search_convert = jira.search_issues(jqlfilter, maxResults=100)
    milestone_total = issue_search_convert.total
    if(milestone_total != 0) :
        
        groupStart = row - archiTotal
        groupend = milestone_total + row - 1
        milestone_count = 1

        print("groupstart: {}, grpEnd: {}".format(groupStart, groupend))
        startpt = startSP+13
        prior_startpt = 0
        prior_milestone_release = 0
        for key in issue_search_convert:
            ###########################
            ### get milestones'info ###
            ###########################
            milestone = jira.issue(key)
            milestone_type = milestone.fields.issuetype.name
            milestone_key = milestone.key
            milestone_summary = milestone.fields.summary
            milestone_assignee_id = milestone.fields.assignee.name
            milestone_assignee_row = milestone.fields.assignee.displayName
            m_assignee_temp_list = milestone_assignee_row.split(" ")
            milestone_assignee = m_assignee_temp_list[0]
            milestone_due = milestone.fields.duedate
            milestone_status = milestone.fields.status.name

            ################ due change history ################
            milestone_label = milestone.fields.labels
            result = []
            changeSprintList = []
            for s in milestone_label:
                if("일정" in s):
                    result.append(s)

            for sp in result:
                sprint = sp[7:]
                changeSprintList.append(int(sprint))

            milestone_delay_factor = 0

            if(len(changeSprintList) == 1):
                milestone_delay_factor = 1
            elif(len(changeSprintList) > 1):
                milestone_delay_factor = 2

            print("labels: {}, due history: {}".format(milestone_label, changeSprintList))
            ####################################################
            if(milestone_due == None):
                milestone_release = 1
                msp = "none"
            else:
                milestone_release = sprintConvertor(milestone_due)
                msp = milestone_release

            if(msp == prior_milestone_release):
                startpt = prior_startpt + 1

            mSpColumn = milestone_release + 12
            sheet1.cell(row=row, column=1).value = str(
                initiative_count) + "-M" + str(milestone_count)
            sheet1.cell(row=row, column=2).value = milestone_type
            sheet1.cell(row=row, column=3).value = milestone_key
            sheet1.cell(row=row, column=4).value = milestone_summary
            sheet1.cell(row=row, column=5).value = milestone_assignee
            sheet1.cell(row=row, column=6).value = milestone_status
            sheet1.cell(row=row, column=7).value = milestone_assignee_id
            sheet1.cell(row=row, column=8).value = "=VLOOKUP(" + \
                sheet1.cell(row=row, column=7).coordinate+",org!$A:$D,4,0)"
            milestone_count = milestone_count + 1  # index
            sheet1.cell(row=row, column=9).value = fixver
            sheet1.cell(row=row, column=10).value = prior
            sheet1.cell(row=row, column=11).value = scm
            # jira link
            sheet1.cell(
                row=row, column=3).hyperlink = "http://hlm.lge.com/issue/browse/"+milestone_key
            ##############################
            ### fill the color to cell ###
            ##############################
            # 1. fill Total Bar
            sheet1.cell(row=row, column=mSpColumn).value = msp
            if(startpt - mSpColumn >= 0) :
                for k in range(mSpColumn, mSpColumn+1):
                    sheet1.cell(row=row, column=k).fill = PatternFill(
                        fill_type='solid', fgColor="5F9EA0")
            else :
                for k in range(startpt, mSpColumn+1):
                    sheet1.cell(row=row, column=k).fill = PatternFill(
                        fill_type='solid', fgColor="5F9EA0")

            # 2. marking Changing
            if(milestone_delay_factor != 0):
                for changeSP in changeSprintList:
                    sheet1.cell(row=row, column=changeSP + 12).value = "SP" + str(changeSP)

            if(milestone_delay_factor == 1):
                for k in range(changeSprintList[0]+12, changeSprintList[0]+13):
                    sheet1.cell(row=row, column=k).fill = PatternFill(
                        fill_type='solid', fgColor="5F9EA0")
                for k in range(changeSprintList[0]+13, mSpColumn+1):
                    sheet1.cell(row=row, column=k).fill = PatternFill(
                        fill_type='solid', fgColor="ffb142")

            elif(milestone_delay_factor == 2):
                print("Check!!!! {} vs {}".format(startpt, changeSprintList[0]+13))
                for k in range(changeSprintList[0]+12, changeSprintList[0]+13):
                    sheet1.cell(row=row, column=k).fill = PatternFill(
                        fill_type='solid', fgColor="5F9EA0")
                for k in range(changeSprintList[0]+13, changeSprintList[1]+13):
                    sheet1.cell(row=row, column=k).fill = PatternFill(
                        fill_type='solid', fgColor="ffb142")
                for k in range(changeSprintList[1]+13, mSpColumn+1):
                    sheet1.cell(row=row, column=k).fill = PatternFill(
                        fill_type='solid', fgColor="DC143C")

            print("row:{}, type:{}, key: {}, stpt: {}, mspcol: {}".format(
                row, milestone_type, milestone_key, startpt, mSpColumn))
            row = row+1
            prior_startpt = startpt
            prior_milestone_release = milestone_release
            startpt = mSpColumn + 1

        sheet1.row_dimensions.group(start=groupStart, end=groupend)

    return row



def trigger(issue_search_convert, sheetname):
    # create Sheet
    sheet1 = wb.create_sheet(title=sheetname)
    # 1st row(default)
    defaultSheet = ["idx.", "Type", "Key", "Summary", "Owner","Status",  "ID", "Team", "FixVer.", "PMO", "SCM", "risky", "SP1", "SP2", "SP3", "SP4", "SP5", "SP6", "SP7", "SP8", "SP9", "SP10", "SP11", "SP12",
                    "SP13", "SP14", "SP15", "SP16", "SP17", "SP18", "SP19", "SP20", "SP21", "SP22", "SP23", "SP24", "SP25", "SP26", "SP27", "SP28", "SP29", "SP30", "SP31", "SP32", "SP33", "SP34", "SP35"]
    sheet1.append(defaultSheet)
    row = 2
    initiative_count = 1
    print("### Initiative Info. ###")
    # Create New Jira Tickets
    for key in issue_search_convert:

        issue = jira.issue(key)
        issue_type = issue.fields.issuetype.name
        issue_key = issue.key
        print(issue_key)
        issue_summary = issue.fields.summary
        issue_status = issue.fields.status.name

        if(issue.fields.assignee.name != None) : 
            issue_assignee_row = issue.fields.assignee.displayName
            assignee_temp_list = issue_assignee_row.split(" ")
            issue_assignee = assignee_temp_list[0]
            issue_assignee_id = issue.fields.assignee.name

        else : 
            issue_assignee = "unassigned"
            issue_assignee_id = "unassigned"

        issue_fixver = issue.fields.fixVersions[0].name
        issue_labels = issue.fields.labels
        prior = labelCheck(issue_labels)
        issue_compon = issue.fields.components
        scm = componCheck(issue_compon)

        issue_due = issue.fields.duedate
        if(issue_due == None):
            issue_release = 1
        else:
            issue_release = sprintConvertor(issue_due)

        spColumn = issue_release + 12
        sheet1.cell(row=row, column=1).value = initiative_count
        sheet1.cell(row=row, column=2).value = issue_type
        sheet1.cell(row=row, column=3).value = issue_key
        sheet1.cell(row=row, column=4).value = issue_summary
        sheet1.cell(row=row, column=5).value = issue_assignee
        sheet1.cell(row=row, column=6).value = issue_status
        sheet1.cell(row=row, column=7).value = issue_assignee_id
        sheet1.cell(row=row, column=8).value = "=VLOOKUP(" + \
            sheet1.cell(row=row, column=7).coordinate+",org!$A:$D,4,0)"
        issue_fixver = issue.fields.fixVersions[0].name
        sheet1.cell(row=row, column=9).value = issue_fixver
        sheet1.cell(row=row, column=10).value = prior
        sheet1.cell(row=row, column=11).value = scm

        sheet1.cell(
            row=row, column=3).hyperlink = "http://hlm.lge.com/issue/browse/"+issue_key

        createDate = issue.fields.created
        createDateList = createDate.split("T")
        creatSP = sprintConvertor(createDateList[0])
        thisSP = nowSprintConvertor()

        if(issue_status == "Approved" or issue_status == "BACKLOG REFINEMENT" or issue_status == "READY" or issue_status == "In Progress" or issue_status == "Delivered"):

            histories = issue.changelog.histories

            for history in histories:
                for item in history.items:
                    if(item.toString == "Approved"):
                        apprDate = history.created
                        apprDateList = apprDate.split("T")
                        startSP = sprintConvertor(apprDateList[0])
                    if(item.toString == "ELT REVIEW"):
                        eltDate = history.created
                        eltDateList = eltDate.split("T")
                        eltSP = sprintConvertor(eltDateList[0])
            
            sheet1.cell(row=row, column=spColumn).value = "O"
            for j in range(1, 7):
                sheet1.cell(row=row, column=j).fill = PatternFill(
                    fill_type='solid', fgColor="98FB98")

            for e in range(creatSP+ 13, eltSP+ 13):
                sheet1.cell(row=row, column=e).fill = PatternFill(
                    fill_type='solid', fgColor="FFDAB9")

            for e in range(eltSP+ 13, startSP+ 13):
                sheet1.cell(row=row, column=e).fill = PatternFill(
                    fill_type='solid', fgColor="DCDCDC")

            for k in range(startSP+ 13, spColumn+1):
                sheet1.cell(row=row, column=k).fill = PatternFill(
                    fill_type='solid', fgColor="000000")

            print("row:{}, index:{}, type:{}, key: {}, start: {}".format(
                row, initiative_count, issue_type, issue_key, startSP))

            
            row = row + 1

            row, archiCount = archiListUp(
                initiative_count, issue_key, issue_fixver, row, startSP, sheet1, prior, scm)
            row = milestoneListUp(initiative_count, issue_key, issue_fixver,
                                row, startSP, sheet1, archiCount,prior, scm)
            initiative_count = initiative_count + 1

        elif(issue_status == "ELT REVIEW"):

            histories = issue.changelog.histories

            for history in histories:
                for item in history.items:
                    if(item.toString == "ELT REVIEW"):
                        eltDate = history.created
                        eltDateList = eltDate.split("T")
                        eltSP = sprintConvertor(eltDateList[0])

            sheet1.cell(row=row, column=spColumn).value = "O"
            for j in range(1, 7):
                sheet1.cell(row=row, column=j).fill = PatternFill(
                    fill_type='solid', fgColor="98FB98")

            for e in range(creatSP+ 13, eltSP+ 13):
                sheet1.cell(row=row, column=e).fill = PatternFill(
                    fill_type='solid', fgColor="FFDAB9")

            for k in range(eltSP+ 13, thisSP+13):
                sheet1.cell(row=row, column=k).fill = PatternFill(
                    fill_type='solid', fgColor="DCDCDC")


            print(issue_status, creatSP, eltSP, thisSP)
            


            row = row + 1
            initiative_count = initiative_count + 1
            print("row:{}, index:{}, type:{}, key: {}".format(
                row, initiative_count, issue_type, issue_key))            

        else : 
            sheet1.cell(row=row, column=spColumn).value = "O"
            for j in range(1, 7):
                sheet1.cell(row=row, column=j).fill = PatternFill(
                    fill_type='solid', fgColor="98FB98")

            for k in range(13, spColumn+1):
                sheet1.cell(row=row, column=k).fill = PatternFill(
                    fill_type='solid', fgColor="000000")

            row = row + 1
            initiative_count = initiative_count + 1
            print("row:{}, index:{}, type:{}, key: {}".format(
                row, initiative_count, issue_type, issue_key))            



    sheet1.column_dimensions['A'].width = 6
    sheet1.column_dimensions['B'].width = 10
    sheet1.column_dimensions['C'].width = 13
    sheet1.column_dimensions['D'].width = 55
    sheet1.column_dimensions['E'].width = 6.5
    sheet1.column_dimensions['H'].width = 25
    sheet1.column_dimensions['G'].hidden = True
    sheet1.column_dimensions["A"].alignment = Alignment(
        horizontal='center', vertical='center')
    sheet1.column_dimensions['B'].alignment = Alignment(
        horizontal='center', vertical='center')
    sheet1.column_dimensions['C'].alignment = Alignment(
        horizontal='center', vertical='center')
    sheet1.column_dimensions['E'].alignment = Alignment(
        horizontal='center', vertical='center')
    sheet1.column_dimensions['G'].alignment = Alignment(
        horizontal='center', vertical='center')
    sheet1.freeze_panes = 'J2'  # 틀고정

    wb.save("D:/test/(200320)webOS SW Gantt.xlsx")
    # wb.save("D:/test/gantttest.xlsx")

if __name__ == "__main__":
    # jira Handle open
    jira = JIRA(DevTracker, basic_auth=(ID, PASSWORD))
    # # Test Filter
    # issue_search_convert1 = jira.search_issues('key in (TVPLAT-50838, TVPLAT-51383)', maxResults=1000, expand="changelog")
    # sheet1 = "test"
    #####################
    ###### Filter #######
    #####################
    # w5.0 mr major #
    issue_search_convert2 = jira.search_issues(
        'filter=01_initiative_total_pm_opened and  fixversion in (\"webOS TV 5.0 MR major") ORDER BY status ASC, due ASC', maxResults=1000, expand="changelog")
    # w6.0 initial(inclusive platform) #
    issue_search_convert1 = jira.search_issues('filter=01_initiative_total_pm_opened and fixversion in (\"webOS TV 6.0 Initial\",\"webOS TV 6.0 Platform\") and assignee is not EMPTY ORDER BY status ASC, due ASC', maxResults=1000, expand="changelog")
    # PMO.0305 #
    issue_search_convert3 = jira.search_issues(
        'labels = PMO.0305', maxResults=1000, expand="changelog")
    sheet2 = "W5.0_MR_major_0320"
    sheet1 = "W6.0_initial_0320"
    wb = openpyxl.load_workbook("D:/test/Gantt.xlsx")
    orgSheet = wb["org"]

    trigger(issue_search_convert1,sheet1)
    trigger(issue_search_convert2,sheet2)
